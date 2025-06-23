import os
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib import messages
from multiprocessing import Pool, cpu_count
from .dict import USUARIOS
from openpyxl import load_workbook
# Bibliotecas da OCC (OpenCascade) para ler arquivos STEP e renderizar imagens
from OCC.Core.STEPControl import STEPControl_Reader
from OCC.Core.IFSelect import IFSelect_RetDone
from OCC.Display.OCCViewer import OffscreenRenderer

# Variaveis globais
caminho_base = r'media\modelos'
caminho_arquivo = r'media\moldes.xlsx'

# Função responsável por gerar a imagem (preview) a partir de um arquivo .stp
def gerar_preview_stp_para_png(caminho_stp, caminho_png):
    # Cria o leitor para arquivos STEP
    reader = STEPControl_Reader()
    status = reader.ReadFile(caminho_stp)
    
    # Verifica se o arquivo foi lido com sucesso
    if status != IFSelect_RetDone:
        raise Exception(f"Erro ao ler o arquivo STEP: {caminho_stp}")
    
    # Transfere o conteúdo do arquivo STEP para a variável shape (forma 3D)
    reader.TransferRoot()
    shape = reader.Shape()

    # Inicializa o renderizador offscreen (sem janela gráfica)
    renderer = OffscreenRenderer(screen_size=(640, 480))

    # Exibe o modelo 3D e salva a imagem no caminho indicado
    renderer.DisplayShape(
        shapes=shape,
        transparency=0.7,
        dump_image=True,
        dump_image_path=os.path.dirname(caminho_png),
        dump_image_filename=os.path.basename(caminho_png)
    )

# Função chamada em paralelo para cada arquivo STEP
def processar_preview(arquivo_info):
    # Desempacota os dados
    pasta_modelos, pasta_previews, arquivo = arquivo_info

    # Define os caminhos para o arquivo STEP e a imagem JPEG correspondente
    caminho_stp = os.path.join(pasta_modelos, arquivo)
    nome_imagem = arquivo.replace('.stp', '.jpeg')
    caminho_png = os.path.join(pasta_previews, nome_imagem)

    try:
        # Só gera nova imagem se não existir ou se o arquivo .stp foi modificado mais recentemente
        if not os.path.exists(caminho_png) or os.path.getmtime(caminho_stp) > os.path.getmtime(caminho_png):
            gerar_preview_stp_para_png(caminho_stp, caminho_png)

        # Retorna um dicionário com as informações que serão exibidas no HTML
        return {
            'nome': arquivo,
            'imagem_url': f"/media/modelos/previews/{nome_imagem}"
        }

    except Exception as e:
        # Caso algo dê errado, mostra o erro no terminal e retorna None
        print(f"Erro ao gerar preview para {arquivo}: {e}")
        return None

# View do Django que lista os modelos e seus previews
def listar_modelos_step(request):
    # Define o caminho para a pasta com os modelos (.stp)
    pasta_modelos = os.path.join(settings.MEDIA_ROOT, 'modelos')
    
    # Pasta onde serão salvas as imagens (previews)
    pasta_previews = os.path.join(pasta_modelos, 'previews')
    os.makedirs(pasta_previews, exist_ok=True)  # Cria se não existir

    # Lista todos os arquivos .stp na pasta
    arquivos_stp = [f for f in os.listdir(pasta_modelos) if f.lower().endswith('.stp')]

    # Prepara os dados que serão enviados para cada processo (arquivo + pastas)
    args = [(pasta_modelos, pasta_previews, arquivo) for arquivo in arquivos_stp]

    # Cria um pool de processos para acelerar o processamento usando múltiplos núcleos
    with Pool(processes=cpu_count()) as pool:
        resultados = pool.map(processar_preview, args)

    # Remove da lista os arquivos que falharam (None)
    previews = [r for r in resultados if r]

    # Renderiza a página HTML com as imagens geradas
    return render(request, 'visualizador/preview.html', {'previews': previews})

# Metodo de Login
def login(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').lower().strip()
        password = request.POST.get('password', '').strip()

        dados = USUARIOS.get(username)

        if dados and password == dados['senha']:
            request.session['usuario'] = username
            request.session['permissao'] = dados['permissao']
            return redirect('pagina_principal')

        return render(request, 'login/login.html', {'erro': 'Usuário ou senha inválidos'})

    return render(request, 'login/login.html')



def pagina_principal(request, nome_aba=None):
    elementos = []
    dados = []
    mensagem = None
    abas_excel = []

    # Lista todas as pastas do diretório
    try:
        for nome in os.listdir(caminho_base):
            caminho_completo = os.path.join(caminho_base, nome)
            if os.path.isdir(caminho_completo):
                elementos.append({'tipo': 'pasta', 'nome': nome})
    except Exception as e:
        elementos = [{'tipo': 'erro', 'nome': f"Erro: {str(e)}"}]

    # Verifica e carrega abas do Excel
    if os.path.exists(caminho_arquivo):
        try:
            wb = load_workbook(caminho_arquivo, data_only=True)
            abas_excel = wb.sheetnames

            if nome_aba:
                if nome_aba in abas_excel:
                    aba = wb[nome_aba]
                        # Itera pelas linhas a partir da 1ª (ou 2ª se quiser pular cabeçalho)
                    for row in aba.iter_rows(min_row=2, max_col=11):  
                        linha_num = row[0].row  # pega o número da linha
                        item = row[1].value
                        chegada_aco = bool(row[4].value)
                        programa = bool(row[4].value)
                        maquina_1 = bool(row[5].value)
                        maquina_2 = bool(row[6].value)
                        maquina_3 = bool(row[7].value)
                        maquina_4 = bool(row[8].value)
                        maquina_5 = bool(row[9].value)  
                        maquina_6 = bool(row[10].value)
                        em_maquina = any([
                            maquina_1, maquina_2, maquina_3, 
                            maquina_4, maquina_5, maquina_6
                        ])
                        # Força valores booleanos para evitar erros
                        chegada_aco = bool(row[3].value)
                        programa = bool(row[4].value)

                        # Lógica para determinar o status customizado
                        if chegada_aco and programa and not em_maquina:
                            status_custom = "Pronto para usinar"
                        elif chegada_aco and programa and em_maquina:
                            status_custom = "Usinando"
                        elif chegada_aco and not programa:
                            status_custom = "Aguardando programa"
                        elif programa and not chegada_aco:
                            status_custom = "Aguardando aço"
                        else:
                            status_custom = "Aguardando aço e programa"
                        dados.append({
                            'linha': linha_num,
                            'item': item, 
                            'chegada_aco': bool(chegada_aco), 
                            'programa': bool(programa),
                            'status_custom': status_custom
                        })
                else:
                    mensagem = f"Não apresenta dados do molde '{nome_aba}'."
        except Exception as e:
            mensagem = f"Erro ao abrir o Excel: {e}"
    else:
        mensagem = "Arquivo moldes.xlsx não encontrado."

    return render(request, 'tela_principal/principal.html', {
        'elementos': elementos,
        'dados': dados,
        'aba': nome_aba,
        'mensagem': mensagem,
        'permissao': request.session.get('permissao')
    })

def atualizar_status(request):
    if request.method == 'POST':
        nome_aba = request.POST.get('nome_aba')
        wb = load_workbook(caminho_arquivo)
        if nome_aba and nome_aba in wb.sheetnames:
            ws = wb[nome_aba]  # Atualiza na aba correta
        else:
            ws = wb.active  # Fallback se algo der errado
            
        for row in ws.iter_rows(min_row=2, max_col=6):  # base para pegar linha
            linha_num = row[0].row
            #Campos e colunas de cada item
            campos = {
                'chegada_aco': 4,
                'programa': 5,
            }
            #Um for para inserir no excel a informação 
            for campo_nome, col_num in campos.items():
                campo_form = f'{campo_nome}_{linha_num}'
                valor_checkbox = '☑' if request.POST.get(campo_form) == 'on' else ''
                ws.cell(row=linha_num, column=col_num).value = valor_checkbox

        # Salva em um caminho
        wb.save(caminho_arquivo)

        messages.success(request, 'Mudanças salvas com sucesso.')
        return redirect(f'/pagina/{nome_aba}')


def checklist(request, nome_aba=None):
    elementos = []
    dados = []
    mensagem = None
    abas_excel = []

    # Lista todas as pastas do diretório
    try:
        for nome in os.listdir(caminho_base):
            caminho_completo = os.path.join(caminho_base, nome)
            if os.path.isdir(caminho_completo):
                elementos.append({'tipo': 'pasta', 'nome': nome})
    except Exception as e:
        elementos = [{'tipo': 'erro', 'nome': f"Erro: {str(e)}"}]

    # Verifica e carrega abas do Excel
    if os.path.exists(caminho_arquivo):
        try:
            wb = load_workbook(caminho_arquivo, data_only=True)
            abas_excel = wb.sheetnames

            if nome_aba:
                if nome_aba in abas_excel:
                    aba = wb[nome_aba]
                        # Itera pelas linhas a partir da 1ª (ou 2ª se quiser pular cabeçalho)
                    for row in aba.iter_rows(min_row=2, max_col=11):  
                        linha_num = row[0].row  # pega o número da linha
                        item = row[1].value
                        chegada_aco = bool(row[4].value)
                        programa = bool(row[4].value)
                        maquina_1 = bool(row[5].value)
                        maquina_2 = bool(row[6].value)
                        maquina_3 = bool(row[7].value)
                        maquina_4 = bool(row[8].value)
                        maquina_5 = bool(row[9].value)  
                        maquina_6 = bool(row[10].value)
                        em_maquina = any([
                            maquina_1, maquina_2, maquina_3, 
                            maquina_4, maquina_5, maquina_6
                        ])
                        # Força valores booleanos para evitar erros
                        chegada_aco = bool(row[3].value)
                        programa = bool(row[4].value)

                        # Lógica para determinar o status customizado
                        if chegada_aco and programa and not em_maquina:
                            status_custom = "Pronto para usinar"
                        elif chegada_aco and programa and em_maquina:
                            status_custom = "Usinando"
                        elif chegada_aco and not programa:
                            status_custom = "Aguardando programa"
                        elif programa and not chegada_aco:
                            status_custom = "Aguardando aço"
                        else:
                            status_custom = "Indefinido"
                        dados.append({
                            'linha': linha_num,
                            'item': item,
                            'maquina_1': bool(maquina_1),  
                            'maquina_2': bool(maquina_2),  
                            'maquina_3': bool(maquina_3),  
                            'maquina_4': bool(maquina_4),  
                            'maquina_5': bool(maquina_5),  
                            'maquina_6': bool(maquina_6),  
                            'status_custom': status_custom
                        })
                        print(status_custom)
                else:
                    mensagem = f"Não apresenta dados do molde '{nome_aba}'."
        except Exception as e:
            mensagem = f"Erro ao abrir o Excel: {e}"
    else:
        mensagem = "Arquivo moldes.xlsx não encontrado."

    return render(request, 'checklist/checklist.html', {
        'elementos': elementos,
        'dados': dados,
        'aba': nome_aba,
        'mensagem': mensagem
    })

def atualizar_status_checklist(request):
    if request.method == 'POST':
        nome_aba = request.POST.get('nome_aba')
        wb = load_workbook(caminho_arquivo)
        if nome_aba and nome_aba in wb.sheetnames:
            ws = wb[nome_aba]  # Atualiza na aba correta
        else:
            ws = wb.active  # Fallback se algo der errado
            
        for row in ws.iter_rows(min_row=2, max_col=11):  # base para pegar linha
            linha_num = row[0].row
            #Campos e colunas de cada item
            campos = {
                'maquina_1': 6,
                'maquina_2': 7,
                'maquina_3': 8,
                'maquina_4': 9,
                'maquina_5': 10,
                'maquina_6': 11,
            }
            #Um for para inserir no excel a informação 
            for campo_nome, col_num in campos.items():
                campo_form = f'{campo_nome}_{linha_num}'
                print(campo_form)
                valor_checkbox = '☑' if request.POST.get(campo_form) == 'on' else ''
                ws.cell(row=linha_num, column=col_num).value = valor_checkbox

        # Salva em um novo caminho se quiser manter cópia
        wb.save(caminho_arquivo)
        
    messages.success(request, 'Mudanças salvas com sucesso.')
    return redirect(f'/checklist/{nome_aba}')